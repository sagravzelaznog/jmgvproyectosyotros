import pulp
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

groups_data = {
    "1A": [
        ("LENGUA Y COMUNICACION I", "LIC. ALEJANDRA RIVAS", 3),
        ("CIENCIAS SOCIALES I", "LIC. ALEJANDRA RIVAS", 2),
        ("HUMANIDADES I", "LIC. ALEJANDRA RIVAS", 4),
        ("INGLES 1", "DRA. IVONE ORTEGA REYES", 3),
        ("PENSAMIENTO MATEMATICO I", "M.C. MANUEL GONZALEZ", 4),
        ("LA MATERIA Y SUS INTERACCIONES", "DRA. IVONE ORTEGA REYES", 4),
        ("CULTURA DIGITAL I", "M.C. KARLA ORTEGA REYES", 3),
        ("LABORATORIO DE INVESTIGACION I", "M.C. KARLA ORTEGA REYES", 3),
        ("ELECTRICIDAD", "M.C. MANUEL GONZALEZ", 1),
        ("COMPUTACION", "M.C. KARLA ORTEGA REYES", 1),
        ("ASISTENTE EDUC.", "M.C. KARLA ORTEGA REYES", 1),
        ("ECOLOGIA NUT", "DRA. IVONE ORTEGA REYES", 1),
        ("ECOLOGIA AMBIENTAL", "DRA. IVONE ORTEGA REYES", 1)
    ],
    "1B": [
        ("LENGUA Y COMUNICACION I", "LIC. ALEJANDRA RIVAS", 3),
        ("CIENCIAS SOCIALES I", "LIC. ALEJANDRA RIVAS", 2),
        ("HUMANIDADES I", "LIC. ALEJANDRA RIVAS", 4),
        ("INGLES 1", "DRA. IVONE ORTEGA REYES", 3),
        ("PENSAMIENTO MATEMATICO I", "M.C. MANUEL GONZALEZ", 4),
        ("LA MATERIA Y SUS INTERACCIONES", "DRA. IVONE ORTEGA REYES", 4),
        ("CULTURA DIGITAL I", "M.C. KARLA ORTEGA REYES", 3),
        ("LABORATORIO DE INVESTIGACION I", "M.C. KARLA ORTEGA REYES", 3),
        ("ELECTRICIDAD", "M.C. MANUEL GONZALEZ", 1),
        ("COMPUTACION", "M.C. KARLA ORTEGA REYES", 1),
        ("ASISTENTE EDUC.", "M.C. KARLA ORTEGA REYES", 1),
        ("ECOLOGIA NUT", "DRA. IVONE ORTEGA REYES", 1),
        ("ECOLOGIA AMBIENTAL", "DRA. IVONE ORTEGA REYES", 1)
    ],
    "3ro": [
        ("ECOSISTEMAS: INTERACC, ENERGIA Y DINAM.", "DRA. IVONE ORTEGA REYES", 4),
        ("HUMANIDADES III", "LIC. ALEJANDRA RIVAS", 5),
        ("PENSAMIENTO MATEMATICO III", "M.C. MANUEL GONZALEZ", 4),
        ("INGLES III", "DRA. IVONE ORTEGA REYES", 3),
        ("LENGUA Y COMUNICACIÓN III", "LIC. ALEJANDRA RIVAS", 3),
        ("TALLER DE CIENCIAS II", "M.C. KARLA ORTEGA REYES", 3),
        ("INFORMATICA III", "M.C. MANUEL GONZALEZ", 4),
        ("ASISTENTE EDUC.", "M.C. KARLA ORTEGA REYES", 1),
        ("COMPUTACION III", "M.C. MANUEL GONZALEZ", 1),
        ("ELECTRICIDAD III", "M.C. MANUEL GONZALEZ", 1),
        ("ECOLOGIA NUT.", "DRA. IVONE ORTEGA REYES", 1),
        ("ECOLOGIA AMB.", "DRA. IVONE ORTEGA REYES", 1)
    ],
    "5to": [
        ("LA ENERGIA EN LOS PROCESOS DE LA VIDA DIARIA", "M.C. KARLA ORTEGA REYES", 4),
        ("CONCIENCIA HISTORICA II", "LIC. ALEJANDRA RIVAS", 3),
        ("TALLER DE PROB. Y ESTAD. 1", "DRA. IVONE ORTEGA REYES", 3),
        ("INGLES 5", "DRA. IVONE ORTEGA REYES", 3),
        ("ANALISIS DE FENOMENOS FISICOS", "M.C. MANUEL GONZALEZ", 3),
        ("SALUD INTEGRAL I", "M.C. KARLA ORTEGA REYES", 3),
        ("INFORMATICA 5", "M.C. MANUEL GONZALEZ", 4),
        ("ASISTENTE EDUC.", "M.C. KARLA ORTEGA REYES", 1),
        ("ECOLOGIA NUTRIC", "M.C. KARLA ORTEGA REYES", 1),
        ("TALLER DE PENSAMIENTO VARIACIONAL 1", "M.C. MANUEL GONZALEZ", 3),
        ("ELECTRICIDAD", "M.C. MANUEL GONZALEZ", 1),
        ("COMPUTACION", "M.C. MANUEL GONZALEZ", 1)
    ]
}

def t(h, m): return h * 60 + m

slots = {
    "1A": [
        (t(8, 0), t(8, 45)),
        (t(8, 45), t(9, 30)),
        (t(9, 30), t(10, 15)),
        (t(10, 15), t(10, 55)),
        (t(11, 15), t(12, 0)),
        (t(12, 0), t(12, 45)),
        (t(12, 45), t(13, 30)),
        (t(13, 30), t(14, 15)),
    ],
    "1B": [
        (t(8, 0), t(8, 45)),
        (t(8, 45), t(9, 30)),
        (t(9, 30), t(10, 20)),
        (t(10, 40), t(11, 25)),
        (t(11, 25), t(12, 10)),
        (t(12, 10), t(12, 55)),
        (t(12, 55), t(13, 40)),
        (t(13, 40), t(14, 25)),
    ],
    "3ro": [
        (t(7, 15), t(8, 0)),
        (t(8, 0), t(8, 45)),
        (t(8, 45), t(9, 30)),
        (t(9, 30), t(10, 0)),
        (t(10, 20), t(11, 5)),
        (t(11, 5), t(11, 50)),
        (t(11, 50), t(12, 35)),
        (t(12, 35), t(13, 20)),
    ],
    "5to": [
        (t(7, 15), t(8, 0)),
        (t(8, 0), t(8, 45)),
        (t(8, 45), t(9, 30)),
        (t(9, 50), t(10, 35)),
        (t(10, 35), t(11, 20)),
        (t(11, 20), t(12, 5)),
        (t(12, 5), t(12, 50)),
        (t(12, 50), t(13, 35)),
    ]
}

def overlap(s1, e1, s2, e2):
    return not (e1 <= s2 or e2 <= s1)

prob = pulp.LpProblem("Horarios", pulp.LpMaximize)

# x[g, d, s, c]
x = {}
days = list(range(5))

for g, classes in groups_data.items():
    for d in days:
        for s_idx in range(len(slots[g])):
            for c_idx, _ in enumerate(classes):
                x[(g, d, s_idx, c_idx)] = pulp.LpVariable(f"x_{g}_{d}_{s_idx}_{c_idx}", cat="Binary")

# 1. Total sessions constraint
for g, classes in groups_data.items():
    for c_idx, (name, teacher, count) in enumerate(classes):
        prob += pulp.lpSum([x[(g, d, s_idx, c_idx)] for d in days for s_idx in range(len(slots[g]))]) == count

# 2. At most one class per slot per group
for g in groups_data.keys():
    for d in days:
        for s_idx in range(len(slots[g])):
            prob += pulp.lpSum([x[(g, d, s_idx, c_idx)] for c_idx in range(len(groups_data[g]))]) <= 1

# 3. No double sessions on the same day
for g, classes in groups_data.items():
    for c_idx in range(len(classes)):
        for d in days:
            prob += pulp.lpSum([x[(g, d, s_idx, c_idx)] for s_idx in range(len(slots[g]))]) <= 1

# 4. Teacher Unavailability
for g, classes in groups_data.items():
    for c_idx, (name, teacher, count) in enumerate(classes):
        for d in days:
            for s_idx, (st, et) in enumerate(slots[g]):
                if teacher == "M.C. KARLA ORTEGA REYES":
                    if overlap(st, et, t(8, 45), t(11, 15)):
                        prob += x[(g, d, s_idx, c_idx)] == 0
                if teacher == "M.C. MANUEL GONZALEZ":
                    if overlap(st, et, t(12, 45), t(13, 30)):
                        prob += x[(g, d, s_idx, c_idx)] == 0

# 5. Teacher Overlap (Cannot teach two classes at the same time)
teachers = ["LIC. ALEJANDRA RIVAS", "DRA. IVONE ORTEGA REYES", "M.C. MANUEL GONZALEZ", "M.C. KARLA ORTEGA REYES"]
for d in days:
    for teacher in teachers:
        # Get all slots across all groups for this teacher
        teacher_slots = []
        for g, classes in groups_data.items():
            for c_idx, (name, t_name, count) in enumerate(classes):
                if t_name == teacher:
                    for s_idx, (st, et) in enumerate(slots[g]):
                        teacher_slots.append((g, c_idx, s_idx, st, et))
        
        # Check all pairs
        for i in range(len(teacher_slots)):
            for j in range(i + 1, len(teacher_slots)):
                g1, c1, s1, st1, et1 = teacher_slots[i]
                g2, c2, s2, st2, et2 = teacher_slots[j]
                if (g1 != g2 or s1 != s2): # If same group and same slot, already handled by Constraint 2
                    if overlap(st1, et1, st2, et2):
                        prob += x[(g1, d, s1, c1)] + x[(g2, d, s2, c2)] <= 1

# 6. Objective: pack classes early
obj = []
for g in groups_data.keys():
    for d in days:
        for s_idx in range(len(slots[g])):
            weight = 100 - s_idx * 10
            for c_idx in range(len(groups_data[g])):
                obj.append(weight * x[(g, d, s_idx, c_idx)])

prob += pulp.lpSum(obj)

status = prob.solve(pulp.PULP_CBC_CMD(msg=False, timeLimit=60))
print("Status:", pulp.LpStatus[status])

if status != pulp.LpStatusOptimal and status != pulp.LpStatusFeasible:
    print("Could not find a feasible schedule!")
    exit(1)

# Generate Excel
wb = Workbook()
wb.remove(wb.active)

recesses = {
    "1A": "10:55 - 11:15",
    "1B": "10:20 - 10:40",
    "3ro": "10:00 - 10:20",
    "5to": "9:30 - 9:50"
}

for g in ["1A", "1B", "3ro", "5to"]:
    ws = wb.create_sheet(title=g)
    
    # Headers
    days_str = ["LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES"]
    ws.cell(row=1, column=1, value="HORARIO")
    for idx, d in enumerate(days_str):
        ws.cell(row=1, column=2+idx, value=d)
    
    # Fill times
    row_idx = 2
    for s_idx, (st, et) in enumerate(slots[g]):
        start_str = f"{st//60:02d}:{st%60:02d}"
        end_str = f"{et//60:02d}:{et%60:02d}"
        ws.cell(row=row_idx, column=1, value=f"{start_str} - {end_str}")
        row_idx += 1
        
    # Insert recess conceptually in output
    # Let's map back classes
    for d in days:
        for s_idx in range(len(slots[g])):
            assigned = None
            for c_idx, (name, t_name, _) in enumerate(groups_data[g]):
                if pulp.value(x[(g, d, s_idx, c_idx)]) == 1.0:
                    assigned = f"{name}\n({t_name})"
                    break
            if assigned:
                cell = ws.cell(row=2+s_idx, column=2+d, value=assigned)
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Formatting
    for col in range(1, 7):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    for r in range(1, len(slots[g]) + 2):
        ws.row_dimensions[r].height = 50

# Save
wb.save(r"c:\Users\admin\Documents\000 A PREPA\planeaciones especialidades\Proyectos y Otros\HORARIOS\HORARIOS_GENERADOS.xlsx")
print("Saved HORARIOS_GENERADOS.xlsx")
